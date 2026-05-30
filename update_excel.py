# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook('Donnees_projet_complet.xlsx')
df = pd.read_csv('energydata_ready_for_machine_learning.csv', parse_dates=['date'])
df = df.sort_values('date').reset_index(drop=True)

# ── Helpers ──────────────────────────────────────────────────────────────
def hf():   return Font(bold=True, size=10, color='FFFFFF', name='Arial')
def tf():   return Font(bold=True, size=12, color='FFFFFF', name='Arial')
def cf():   return Font(size=9, name='Arial')
def tfill():return PatternFill('solid', start_color='7B2FBE')
def hfill():return PatternFill('solid', start_color='4A1A7A')
def afill():return PatternFill('solid', start_color='F8F4FF')
def wfill():return PatternFill('solid', start_color='FFFFFF')
def ctr():  return Alignment(horizontal='center', vertical='center', wrap_text=True)
def lft():  return Alignment(horizontal='left', vertical='center', wrap_text=True)
def brd():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

# ════════════════════════════════════════════════════════════════════════
# FEUILLE 1 : Description Variables
# ════════════════════════════════════════════════════════════════════════
ws = wb['Description Variables']
# Supprimer les merges existantes avant d'écrire
import copy
for merge in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merge))

ws['A1'] = 'Projet 2 - Prediction de la Consommation Energetique | Dataset Enrichi (RETOUR)'
ws['A1'].font = tf(); ws['A1'].fill = tfill(); ws['A1'].alignment = lft()

ws['A2'] = 'Source : energydata_ready_for_machine_learning.csv - 3098 observations horaires | Jan-Mai 2016 | Residentiel / Commercial / Industriel'
ws['A2'].font = Font(size=10, color='4A1A7A', name='Arial', italic=True)

nouvelles_vars = [
    (1,  'date',             'Date et heure',            'Date',      'YYYY-MM-DD HH:MM', 'Dataset enrichi',     'Present'),
    (2,  'Consommation_kWh', 'Consommation (CIBLE)',      'Numerique', 'kWh',              'Dataset enrichi',     'Present'),
    (3,  'Temperature',      'Temperature exterieure',    'Numerique', 'C',                'UCI agrege horaire',  'Present'),
    (4,  'Humidite',         'Humidite relative',         'Numerique', '%',                'UCI agrege horaire',  'Present'),
    (5,  'Heure',            'Heure de la journee',       'Numerique', '0-23',             'Extrait de date',     'Present'),
    (6,  'Jour_semaine',     'Jour de la semaine',        'Numerique', '1-7',              'Extrait de date',     'Present'),
    (7,  'Jour_ferie',       'Jour ferie',                'Binaire',   '0/1',              'Calendrier 2016',     'Present'),
    (8,  'Type_batiment',    'Type de batiment',          'Categoriel','Res/Com/Ind',      'Simulation enrichie', 'Present'),
    (9,  'Surface',          'Surface du batiment',       'Numerique', 'm2',               'Simulation enrichie', 'Present'),
    (10, 'Occupants',        'Nombre occupants',          'Numerique', 'Personnes',        'Simulation enrichie', 'Present'),
    (11, 'Consommation_J-1', 'Consommation J-1 (lag 24h)','Numerique','kWh',              'Feature engineering', 'Present'),
    (12, 'Consommation_H-1', 'Consommation H-1 (lag 1h)', 'Numerique','kWh',              'Feature engineering', 'Present'),
    (13, 'month',            'Mois',                      'Numerique', '1-12',             'Feature engineering', 'Present'),
    (14, 'is_weekend',       'Weekend',                   'Binaire',   '0/1',              'Feature engineering', 'Present'),
    (15, 'hour_sin/cos',     'Encodage cyclique heure',   'Numerique', '[-1,1]',           'Feature engineering', 'Present'),
    (16, 'dow_sin/cos',      'Encodage cyclique jour',    'Numerique', '[-1,1]',           'Feature engineering', 'Present'),
    (17, 'month_sin/cos',    'Encodage cyclique mois',    'Numerique', '[-1,1]',           'Feature engineering', 'Present'),
    (18, 'conso_H_1/2/6',    'Lags horaires',             'Numerique', 'kWh',              'Feature engineering', 'Present'),
    (19, 'conso_J_1/7',      'Lags journaliers',          'Numerique', 'kWh',              'Feature engineering', 'Present'),
    (20, 'rolling_mean_6h',  'Moyenne mobile 6h',         'Numerique', 'kWh',              'Feature engineering', 'Present'),
    (21, 'rolling_mean_24h', 'Moyenne mobile 24h',        'Numerique', 'kWh',              'Feature engineering', 'Present'),
]

for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = None

headers = ['#', 'Colonne Dataset', 'Variable Projet', 'Type', 'Unite', 'Source', 'Statut']
widths  = [4,   22,                28,                 12,     10,      22,        10]
for c, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = hf(); cell.fill = hfill(); cell.alignment = ctr(); cell.border = brd()
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[4].height = 22

for r, row_data in enumerate(nouvelles_vars, 5):
    fill = afill() if r % 2 == 0 else wfill()
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = cf(); cell.fill = fill
        cell.alignment = ctr() if c == 1 else lft()
        cell.border = brd()
    ws.row_dimensions[r].height = 18

print('Feuille Description Variables OK')

# ════════════════════════════════════════════════════════════════════════
# FEUILLE 2 : Donnees
# ════════════════════════════════════════════════════════════════════════
ws2 = wb[wb.sheetnames[1]]  # 'Données'
for merge in list(ws2.merged_cells.ranges):
    ws2.unmerge_cells(str(merge))
for row in ws2.iter_rows():
    for cell in row:
        if cell.__class__.__name__ != 'MergedCell':
            cell.value = None

cols_show = ['date','Consommation_kWh','Temperature','Humidite','Heure','Jour_semaine',
             'Jour_ferie','Type_batiment','Surface','Occupants',
             'Consommation_J-1','Consommation_H-1','conso_H_1','conso_J_1',
             'rolling_mean_6h','rolling_mean_24h','is_weekend','month']

# Adapter aux noms reels des colonnes
real_cols = {}
for c in cols_show:
    if c in df.columns:
        real_cols[c] = c
    else:
        for dc in df.columns:
            if dc.lower().replace(' ','_') == c.lower().replace(' ','_'):
                real_cols[c] = dc
                break
        if c not in real_cols:
            real_cols[c] = None

cols_available = [c for c in cols_show if real_cols.get(c) in df.columns]
real_names = [real_cols[c] for c in cols_available]

titre = 'Dataset Enrichi - 3098 obs x ' + str(len(cols_available)) + ' variables'
ws2.cell(row=1, column=1).value = titre
ws2.cell(row=1, column=1).font = tf()
ws2.cell(row=1, column=1).fill = tfill()
ws2.cell(row=1, column=1).alignment = lft()
ws2.merge_cells('A1:' + get_column_letter(len(cols_available)) + '1')

for c_idx, (col, real) in enumerate(zip(cols_available, real_names), 1):
    cell = ws2.cell(row=2, column=c_idx, value=col)
    cell.font = hf(); cell.fill = hfill(); cell.alignment = ctr(); cell.border = brd()
    ws2.column_dimensions[get_column_letter(c_idx)].width = max(12, len(col) + 2)
ws2.row_dimensions[2].height = 22

df_show = df[real_names].copy()
df_show[real_names[0]] = df_show[real_names[0]].astype(str)

for r_idx, row_vals in enumerate(df_show.itertuples(index=False), 3):
    fill = afill() if r_idx % 2 == 0 else wfill()
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(size=8, name='Arial')
        cell.fill = fill; cell.border = brd()
        if c_idx > 1 and isinstance(val, float):
            cell.number_format = '0.0000'

print('Feuille Donnees OK (' + str(len(df_show)) + ' lignes)')

# ════════════════════════════════════════════════════════════════════════
# FEUILLE 3 : Statistiques
# ════════════════════════════════════════════════════════════════════════
ws3 = wb[wb.sheetnames[2]]  # 'Statistiques'
for merge in list(ws3.merged_cells.ranges):
    ws3.unmerge_cells(str(merge))
for row in ws3.iter_rows():
    for cell in row:
        cell.value = None

ws3.cell(row=1, column=1).value = 'Statistiques Descriptives - Dataset Enrichi (3098 obs.)'
ws3.cell(row=1, column=1).font = tf()
ws3.cell(row=1, column=1).fill = tfill()
ws3.cell(row=1, column=1).alignment = lft()

num_cols_names = ['Consommation_kWh','Temperature','Humidite','Surface','Occupants',
                  'conso_H_1','conso_J_1','rolling_mean_6h','rolling_mean_24h']
num_cols_real = [c for c in num_cols_names if c in df.columns]
stat_data = df[num_cols_real].describe()

ws3.cell(row=2, column=1, value='Statistique').font = hf()
ws3.cell(row=2, column=1).fill = hfill()
ws3.cell(row=2, column=1).alignment = ctr()
ws3.cell(row=2, column=1).border = brd()
ws3.column_dimensions['A'].width = 14

for c_idx, col in enumerate(num_cols_real, 2):
    cell = ws3.cell(row=2, column=c_idx, value=col)
    cell.font = hf(); cell.fill = hfill(); cell.alignment = ctr(); cell.border = brd()
    ws3.column_dimensions[get_column_letter(c_idx)].width = max(14, len(col) + 2)
ws3.row_dimensions[2].height = 22

for r_idx, stat in enumerate(['count','mean','std','min','25%','50%','75%','max'], 3):
    fill = afill() if r_idx % 2 == 0 else wfill()
    cell = ws3.cell(row=r_idx, column=1, value=stat)
    cell.font = Font(bold=True, size=9, name='Arial')
    cell.fill = fill; cell.border = brd()
    for c_idx, col in enumerate(num_cols_real, 2):
        val = round(float(stat_data.loc[stat, col]), 4)
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.font = cf(); cell.fill = fill; cell.border = brd()
        cell.number_format = '0.0000'

r_note = 12
ws3.cell(row=r_note, column=1, value='Notes :').font = Font(bold=True, size=9, color='7B2FBE', name='Arial')
ws3.cell(row=r_note+1, column=1, value='Type batiment : Residentiel=1910 | Commercial=878 | Industriel=310').font = Font(size=9, italic=True, name='Arial')
ws3.cell(row=r_note+2, column=1, value='Periode : 19 jan. 2016 au 27 mai 2016 | Frequence : horaire | Unite : kWh').font = Font(size=9, italic=True, name='Arial')

print('Feuille Statistiques OK')

# ════════════════════════════════════════════════════════════════════════
# FEUILLE 4 : Repartition -> Resultats Modeles
# ════════════════════════════════════════════════════════════════════════
ws4 = wb[wb.sheetnames[3]]  # 'Répartition'
ws4.title = 'Resultats Modeles'
for merge in list(ws4.merged_cells.ranges):
    ws4.unmerge_cells(str(merge))
for row in ws4.iter_rows():
    for cell in row:
        cell.value = None

ws4.cell(row=1, column=1).value = 'Resultats Comparatifs des Modeles - Dataset Enrichi (3098 obs.)'
ws4.cell(row=1, column=1).font = tf()
ws4.cell(row=1, column=1).fill = tfill()
ws4.cell(row=1, column=1).alignment = lft()
ws4.merge_cells('A1:F1')

mod_headers = ['Modele', 'R2', 'RMSE (kWh)', 'MAE (kWh)', 'MAPE (%)', 'CV R2 moyen']
mod_widths  = [26, 10, 14, 14, 12, 14]
for c_idx, (h, w) in enumerate(zip(mod_headers, mod_widths), 1):
    cell = ws4.cell(row=2, column=c_idx, value=h)
    cell.font = hf(); cell.fill = hfill(); cell.alignment = ctr(); cell.border = brd()
    ws4.column_dimensions[get_column_letter(c_idx)].width = w
ws4.row_dimensions[2].height = 22

modeles = [
    ('Ridge (L2)',          0.4527, 0.2970, 0.1742, 28.37, 0.4063),
    ('Regression Lineaire', 0.4491, 0.2980, 0.1863, 32.63, 0.3929),
    ('Random Forest',       0.3966, 0.3118, 0.1847, 31.35, 0.4083),
    ('Gradient Boosting',   0.3182, 0.3315, 0.1968, 33.63, 0.3456),
]

for r_idx, mod in enumerate(modeles, 3):
    is_best = r_idx == 3
    fill = PatternFill('solid', start_color='EDE0FF') if is_best else (afill() if r_idx % 2 == 0 else wfill())
    for c_idx, val in enumerate(mod, 1):
        cell = ws4.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = fill; cell.border = brd()
        cell.alignment = ctr() if c_idx > 1 else lft()
        cell.font = Font(bold=is_best, size=10 if c_idx==1 else 9, name='Arial',
                         color='4A1A7A' if (is_best and c_idx==1) else ('7B2FBE' if (is_best and c_idx==2) else '000000'))
        if c_idx > 1:
            cell.number_format = '0.00%' if c_idx == 5 else '0.0000'

ws4.cell(row=8, column=1, value='Meilleur modele : Ridge (L2) - R2=0.4527 | CV R2=0.4063').font = Font(bold=True, size=10, color='7B2FBE', name='Arial')
ws4.cell(row=9, column=1, value='Validation : TimeSeriesSplit 5 folds | Split : 80% train / 20% test').font = Font(size=9, italic=True, name='Arial')

print('Feuille Resultats Modeles OK')

wb.save('Donnees_projet_complet.xlsx')
print('Fichier Excel sauvegarde avec succes !')
